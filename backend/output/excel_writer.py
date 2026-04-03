"""
output/excel_writer.py — Excel output from a PackingListRecord matching the structure of Book2.xlsx.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

from config import EXCEL_OUTPUT_DIR
from models import PackingListRecord

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Helper – try import openpyxl once
# ---------------------------------------------------------------------------

def _xl():
    """Lazy import so the module still loads without openpyxl installed."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, Side
        from openpyxl.utils import get_column_letter
        return openpyxl, Alignment, Border, Font, Side, get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl") from exc


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def generate_excel(record: PackingListRecord, job_id: str) -> bytes:
    """
    Build a simple Excel output formatted exactly like Book2.xlsx.
    """
    try:
        openpyxl, Alignment, Border, Font, Side, gcl = _xl()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Packing List"

        # Fonts
        bold_font = Font(bold=True)
        normal_font = Font(bold=False)

        # ── Header Details ──────────────────────────────────────────────────
        
        # Row 1
        ws["C1"] = "Exporter / Manufacturer:"
        ws["C1"].font = bold_font
        ws["D1"] = record.supplier_name.upper() if record.supplier_name != "unknown" else ""

        # Row 3
        ws["A3"] = "Packing list No:"
        ws["A3"].font = bold_font
        ws["B3"] = record.invoice_number or ""

        ws["C3"] = "Date:"
        ws["C3"].font = bold_font
        ws["D3"] = record.delivered_date or ""

        ws["E3"] = "Net Weight:"
        ws["E3"].font = bold_font
        ws["F3"] = record.net_weight or ""

        ws["G3"] = "Total Length:"
        ws["G3"].font = bold_font
        ws["H3"] = record.meters or ""

        # Row 4
        ws["A4"] = "Product:"
        ws["A4"].font = bold_font
        ws["B4"] = record.quality or ""

        # ── Table Headers (Row 8) ───────────────────────────────────────────
        
        headers = [
            "LOT No", 
            "PO #", 
            "Shade", 
            "Roll No", 
            "Length (mts)", 
            "Length (yds)", 
            "Total Points /Roll", 
            "Points / 100m2", 
            "Weight (Gross kgs)", 
            "Weight (Nett kgs)"
        ]

        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col_idx, value=hdr)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            ws.column_dimensions[gcl(col_idx)].width = 15

        # Adjust specific column widths
        ws.column_dimensions["A"].width = 16  # LOT No
        ws.column_dimensions["D"].width = 12  # Roll No
        ws.column_dimensions["G"].width = 18  # Total Points /Roll
        ws.column_dimensions["H"].width = 18  # Points / 100m2
        ws.column_dimensions["I"].width = 18  # Weight (Gross kgs)
        ws.column_dimensions["J"].width = 18  # Weight (Nett kgs)
        ws.column_dimensions["C"].width = 12  # Shade

        # ── Data Rows ───────────────────────────────────────────────────────
        row_idx = 9

        if record.line_items:
            prev_lot = None
            for item in record.line_items:
                # Use explicit yds if available, else calculate
                yds_str = ""
                if item.length_yds is not None:
                    yds_str = str(item.length_yds)
                elif item.meters:
                    yds_str = f"{(float(item.meters) * 1.09361):.2f}"

                lot_val = item.lot or record.lot or ""
                display_lot = lot_val
                if prev_lot and lot_val == prev_lot:
                    display_lot = '"'
                prev_lot = lot_val

                cells = [
                    display_lot,                               # LOT No
                    item.po_number or record.po_number or "",  # PO # (per-roll or doc-level)
                    item.color or record.color or "",          # Shade
                    item.piece_number or "",                   # Roll No
                    item.meters or "",                         # Length (mts)
                    yds_str,                                   # Length (yds)
                    item.points_per_roll or "",                # Total Points /Roll
                    item.points_per_100m2 or "",               # Points / 100m2
                    item.weight_gross_kgs or "",               # Weight (Gross kgs)
                    item.net_weight or ""                      # Weight (Nett kgs)
                ]

                for col_idx, val in enumerate(cells, start=1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.border = thin_border
                    c.alignment = Alignment(horizontal="center")

                row_idx += 1
        else:
            # Document level output if no line items
            yds = (float(record.meters) * 1.09361) if record.meters else None
            yds_str = f"{yds:.2f}" if yds is not None else ""

            cells = [
                record.lot or "",
                record.po_number or "",
                record.color or "",
                record.pieces or "",
                record.meters or "",
                yds_str,
                "",
                "",
                "",
                record.net_weight or ""
            ]
            for col_idx, val in enumerate(cells, start=1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")
            row_idx += 1

        buffer = io.BytesIO()
        wb.save(buffer)
        xlsx_bytes = buffer.getvalue()

        out_path = Path(EXCEL_OUTPUT_DIR) / f"{job_id}.xlsx"
        out_path.write_bytes(xlsx_bytes)
        logger.info("Excel saved to match Book2.xlsx format: %s (%d bytes)", out_path, len(xlsx_bytes))

        return xlsx_bytes

    except Exception as exc:
        logger.error("Excel generation failed for job %s: %s", job_id, exc, exc_info=True)
        raise
